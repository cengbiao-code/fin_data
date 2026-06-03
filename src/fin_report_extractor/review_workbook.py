from __future__ import annotations

import json
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


SUMMARY_COLUMNS = ["field", "value"]

VALIDATION_FAILURE_COLUMNS = [
    "validation_result_id",
    "rule_id",
    "severity",
    "status",
    "message",
    "involved_fact_ids",
    "lhs_value",
    "rhs_value",
    "difference_value",
    "source_pages",
    "suggested_action",
]

RAW_SHEET_NAMES = [
    "balance_sheet_raw",
    "income_statement_raw",
    "cash_flow_raw",
    "notes_revenue_raw",
]

RAW_COLUMNS = [
    "fact_id",
    "raw_table_id",
    "raw_cell_id",
    "extractor_name",
    "page_number",
    "table_index_on_page",
    "row_index",
    "column_index",
    "cell_bbox_json",
    "table_role",
    "statement_scope",
    "period_basis",
    "period_start",
    "period_end",
    "instant_date",
    "raw_label",
    "normalized_concept_id",
    "raw_value",
    "raw_unit",
    "currency",
    "scale_factor",
    "normalized_value",
    "validation_status",
    "review_hint",
]

CORRECTION_COLUMNS = [
    "fact_id",
    "correction_action",
    "corrected_value",
    "corrected_unit",
    "normalized_concept_id",
    "period_basis",
    "statement_scope",
    "table_role",
    "correction_reason",
]

METADATA_COLUMNS = ["key", "value"]

REQUIRED_SHEETS = [
    "summary",
    "validation_failures",
    *RAW_SHEET_NAMES,
    "corrections",
    "_metadata",
]

CORRECTION_ACTIONS = ["confirm", "correct", "remap", "ignore"]

TABLE_ROLE_TO_RAW_SHEET = {
    "statement.balance_sheet": "balance_sheet_raw",
    "statement.income_statement": "income_statement_raw",
    "statement.cash_flow": "cash_flow_raw",
    "note.revenue": "notes_revenue_raw",
}


def export_review_workbook(
    path: str | Path,
    *,
    metadata: dict[str, str],
    failures: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
) -> None:
    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, metadata)
    _write_validation_failures_sheet(wb, failures)
    _write_raw_sheets(wb, raw_rows)
    _write_corrections_sheet(wb)
    _write_metadata_sheet(wb, metadata)

    wb.save(workbook_path)


def read_corrections(path: str | Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "corrections" not in wb.sheetnames:
        raise ValueError("Workbook is missing corrections sheet.")

    ws = wb["corrections"]
    headers = [cell.value for cell in ws[1]]
    missing_columns = [
        column for column in CORRECTION_COLUMNS if column not in headers
    ]
    if missing_columns:
        raise ValueError(f"Corrections sheet is missing columns: {missing_columns}")

    column_indexes = {
        column: headers.index(column)
        for column in CORRECTION_COLUMNS
    }
    rows: list[dict[str, Any]] = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        correction_values = [
            values[column_indexes[column]]
            if column_indexes[column] < len(values)
            else None
            for column in CORRECTION_COLUMNS
        ]
        if all(value is None for value in correction_values):
            continue
        rows.append(dict(zip(CORRECTION_COLUMNS, correction_values)))

    return rows


def _write_summary_sheet(wb: Workbook, metadata: dict[str, str]) -> None:
    ws = wb.create_sheet("summary")
    ws.append(SUMMARY_COLUMNS)
    for key, value in metadata.items():
        ws.append([key, value])
    ws.freeze_panes = "A2"


def _write_validation_failures_sheet(
    wb: Workbook,
    failures: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("validation_failures")
    ws.append(VALIDATION_FAILURE_COLUMNS)
    for row in failures:
        ws.append([
            _excel_value(row.get(column))
            for column in VALIDATION_FAILURE_COLUMNS
        ])
    ws.freeze_panes = "A2"


def _write_raw_sheets(
    wb: Workbook,
    raw_rows: list[dict[str, Any]],
) -> None:
    display_columns = _display_columns(raw_rows)
    raw_columns = [*RAW_COLUMNS, *display_columns]
    worksheets = {
        sheet_name: wb.create_sheet(sheet_name)
        for sheet_name in RAW_SHEET_NAMES
    }

    for ws in worksheets.values():
        ws.append(raw_columns)
        ws.freeze_panes = "A2"

    for row in raw_rows:
        target_sheet = _target_raw_sheet(row)
        if target_sheet not in worksheets:
            continue
        worksheets[target_sheet].append([
            _excel_value(row.get(column))
            for column in raw_columns
        ])


def _write_corrections_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("corrections")
    ws.append(CORRECTION_COLUMNS)
    ws.freeze_panes = "A2"

    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(CORRECTION_ACTIONS)}"',
        allow_blank=True,
    )
    ws.add_data_validation(validation)
    validation.add(f"B2:B1048576")


def _write_metadata_sheet(wb: Workbook, metadata: dict[str, str]) -> None:
    ws = wb.create_sheet("_metadata")
    ws.append(METADATA_COLUMNS)
    for key, value in metadata.items():
        ws.append([key, value])
    ws.freeze_panes = "A2"


def _display_columns(raw_rows: list[dict[str, Any]]) -> list[str]:
    columns = {
        key
        for row in raw_rows
        for key in row
        if key.startswith("display_col_")
    }
    return sorted(columns, key=_display_column_sort_key)


def _display_column_sort_key(column: str) -> tuple[int, str]:
    try:
        return int(column.removeprefix("display_col_")), column
    except ValueError:
        return 10_000, column


def _target_raw_sheet(row: dict[str, Any]) -> str | None:
    target_sheet = row.get("target_sheet")
    if isinstance(target_sheet, str):
        return target_sheet

    table_role = row.get("table_role")
    if isinstance(table_role, str):
        return TABLE_ROLE_TO_RAW_SHEET.get(table_role)

    return None


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value
