from __future__ import annotations

import re
from pathlib import Path
from sqlite3 import Connection
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from fin_report_extractor.fact_extractor import (
    _cells_by_row,
    _source_tables,
    _statement_groups,
)

STATEMENT_METADATA = {
    "statement.balance_sheet": ("balance_sheet", "point_in_time"),
    "statement.income_statement": ("income_statement", "cumulative"),
    "statement.cash_flow": ("cash_flow", "cumulative"),
}

STATEMENT_SHEET_NAMES = {
    "statement.balance_sheet": "资产负债表",
    "statement.income_statement": "利润表",
    "statement.cash_flow": "现金流量表",
}

BS_COLUMNS = ["项目", "期末余额", "期初余额", "来源页"]
IS_CF_COLUMNS = ["项目", "本期发生额", "上期发生额", "来源页"]

REQUIRED_STATEMENT_ROLES = [
    "statement.balance_sheet",
    "statement.income_statement",
    "statement.cash_flow",
]

REQUIRED_LABEL_PATTERNS_BY_MARKET: dict[str, dict[str, list[list[str]]]] = {
    "a_share": {
        "statement.balance_sheet": [
            ["资产总计"],
            ["负债合计"],
            ["所有者权益合计", "股东权益合计", "负债和所有者权益总计"],
        ],
        "statement.income_statement": [
            ["净利润"],
            ["综合收益总额"],
            ["基本每股收益"],
        ],
        "statement.cash_flow": [
            ["经营活动产生的现金流量净额"],
            ["投资活动产生的现金流量净额"],
            ["筹资活动产生的现金流量净额"],
            ["现金及现金等价物净增加额", "期末现金及现金等价物余额"],
        ],
    },
    "hk": {
        "statement.balance_sheet": [
            ["資產總額", "總資產", "total assets"],
            ["負債總額", "總負債", "total liabilities"],
            ["權益總額", "總權益", "total equity"],
        ],
        "statement.income_statement": [
            ["期內盈利", "期內虧損", "年內溢利", "年內虧損", "profit for the period", "profit for the year"],
            ["每股盈利", "每股基本", "每股虧損", "earnings per share", "eps"],
            ["經營盈利", "經營虧損", "operating profit"],
        ],
        "statement.cash_flow": [
            ["經營活動所得現金流量淨額", "經營活動產生的現金流量淨額", "經營活動所用", "net cash generated from operating activities"],
            ["投資活動耗用現金流量淨額", "投資活動所用現金流量淨額", "投資活動所得現金流量淨額", "net cash used in investing activities"],
            ["融資活動", "net cash generated from financing activities"],
            ["期末的現金及現金等價物", "年末現金及現金等價物", "現金及現金等價物增加淨額", "cash and cash equivalents"],
        ],
    },
}

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
LABEL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
AMOUNT_FORMAT = '#,##0.00'
ILLEGAL_EXCEL_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _parse_decimal(raw_value: str | None) -> float | None:
    if raw_value is None:
        return None
    cleaned = raw_value.strip().replace(",", "").replace(" ", "")
    if cleaned in {"", "-", "--"}:
        return None
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    # Only accept pure numeric strings — reject anything with
    # non-numeric characters beyond a leading minus and one decimal point.
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _excel_safe_text(value: str | None) -> str:
    if value is None:
        return ""
    return ILLEGAL_EXCEL_CONTROL_CHARS.sub("", str(value))


def _is_numeric_text(text: str) -> bool:
    return _parse_decimal(text) is not None


def _contains_pattern(text: str, pattern: str) -> bool:
    return pattern.lower() in text.lower()


def _find_label_column(cells: dict[int, dict[str, Any]]) -> int | None:
    for col_index in sorted(cells.keys()):
        raw_text = cells[col_index].get("raw_text")
        if raw_text is None:
            continue
        text = str(raw_text).strip()
        if not text:
            continue
        if _is_numeric_text(text):
            continue
        return col_index
    return None


def _find_value_columns(
    cells: dict[int, dict[str, Any]],
    label_column: int,
) -> tuple[int | None, int | None]:
    numeric_cols = []
    for col_index in sorted(cells.keys()):
        if col_index <= label_column:
            continue
        raw_text = cells[col_index].get("raw_text")
        if raw_text is None:
            continue
        if _is_numeric_text(str(raw_text)):
            numeric_cols.append(col_index)
    current = numeric_cols[0] if len(numeric_cols) > 0 else None
    prior = numeric_cols[1] if len(numeric_cols) > 1 else None
    return current, prior


def _detect_columns(
    header_row: dict[int, dict[str, Any]],
    *,
    statement_type: str,
) -> tuple[int, int | None, int | None]:
    label_col = _find_label_column(header_row)
    if label_col is None:
        label_col = 1

    current_col = None
    prior_col = None

    current_keywords = ["期末", "本期", "current", "period"]
    prior_keywords = ["期初", "上期", "prior", "previous"]

    for col_index in sorted(header_row.keys()):
        if col_index <= label_col:
            continue
        cell = header_row[col_index]
        raw_text = cell.get("raw_text")
        if raw_text is None:
            continue
        text = str(raw_text).strip().lower()
        if any(kw in text for kw in current_keywords):
            current_col = col_index
        elif any(kw in text for kw in prior_keywords):
            prior_col = col_index

    if current_col is None and prior_col is None:
        current_col, prior_col = _find_value_columns(header_row, label_col)

    if current_col is None:
        current_col = label_col + 3 if label_col == 1 else label_col + 1
    if prior_col is None and statement_type == "balance_sheet":
        prior_col = label_col + 6 if label_col == 1 else None

    return label_col, current_col, prior_col


def _is_header_row(cells: dict[int, dict[str, Any]]) -> bool:
    header_text = "\n".join(
        str(cell.get("raw_text") or "")
        for cell in cells.values()
    )
    return any(
        kw in header_text
        for kw in ["项目", "期末", "期初", "本期", "上期", "余额", "发生额", "amount", "period"]
    )


def _adjust_columns_from_data(
    cells: dict[int, dict[str, Any]],
    label_col: int,
    current_col: int,
    prior_col: int | None,
) -> tuple[int, int | None, bool]:
    """Adjust column indices when header text is offset from actual data cells.

    A-share PDF tables often have merged header cells that push the header
    text (like "期末余额") one column to the right of the actual data cells.
    Check whether the detected columns contain numeric text; if not, look
    at neighbouring columns and shift both current and prior together.

    Returns (adjusted_current, adjusted_prior, verified) where *verified*
    indicates whether a numeric data cell was found (so the check consumed
    a real data row, not a section-header row).
    """

    def _has_numeric(col: int | None) -> bool:
        if col is None:
            return False
        text = _cell_text(cells.get(col))
        return text is not None and _is_numeric_text(text)

    if current_col is None:
        return current_col, prior_col, False

    if _has_numeric(current_col):
        return current_col, prior_col, True

    for offset in (-1, 1, -2, 2):
        candidate_current = current_col + offset
        if candidate_current <= label_col:
            continue
        if _has_numeric(candidate_current):
            candidate_prior = (
                prior_col + offset if prior_col is not None else None
            )
            return candidate_current, candidate_prior, True

    return current_col, prior_col, False


def _reconstruct_rows(
    conn: Connection,
    tables: list[Any],
    *,
    statement_type: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    header_label_col: int | None = None
    header_current_col: int | None = None
    header_prior_col: int | None = None
    header_written = False
    columns_verified = False

    for table in tables:
        cells_by_row = _cells_by_row(conn, table.raw_table_id)
        sorted_row_indexes = sorted(cells_by_row.keys())
        # Each table may have a different column layout (e.g. continuation
        # pages can be narrower than the main table).  Re-verify columns
        # from scratch for each new table.
        columns_verified = False

        for row_index in sorted_row_indexes:
            cells = cells_by_row[row_index]

            # Detect header from first table's row 0, or any row that
            # looks like a header.  Continuation tables may start at
            # row 0 with data — only treat row 0 as a header when it
            # carries column-title keywords.
            if row_index == 0 and _is_header_row(cells):
                if not header_written:
                    header_label_col, header_current_col, header_prior_col = _detect_columns(
                        cells, statement_type=statement_type,
                    )
                    rows.append({
                        "label": _cell_text(cells.get(header_label_col)) or "项目",
                        "current": _cell_text(cells.get(header_current_col)) if header_current_col else "",
                        "prior": _cell_text(cells.get(header_prior_col)) if header_prior_col else "",
                        "page": "",
                        "is_header": True,
                    })
                    header_written = True
                continue

            if header_label_col is not None:
                label_col = header_label_col
                current_col = header_current_col
                prior_col = header_prior_col

                # When the header row uses merged cells (common in
                # A-share PDFs), column-title text may sit one column to
                # the right of the actual data values.  Walk forward
                # through data rows until we find one that carries a
                # numeric value — section-header rows like "流动资产："
                # contain no numbers and must be skipped.
                if not columns_verified and current_col is not None:
                    adjusted_current, adjusted_prior, verified = _adjust_columns_from_data(
                        cells, label_col, current_col, prior_col,
                    )
                    header_current_col = adjusted_current
                    header_prior_col = adjusted_prior
                    current_col = header_current_col
                    prior_col = header_prior_col
                    if verified:
                        columns_verified = True

                # Continuation pages may have fewer columns than the
                # main table (e.g. cash flow page 8 has 9 columns but
                # page 9 has only 5).  If the header-detected prior_col
                # doesn't exist in this table, remap it to the rightmost
                # numeric column available.
                if prior_col is not None and prior_col not in cells:
                    available = sorted(
                        c for c in cells.keys()
                        if c > (label_col or 0)
                        and _cell_text(cells.get(c))
                        and _is_numeric_text(_cell_text(cells.get(c)))
                    )
                    if available:
                        prior_col = available[-1]
            else:
                label_col = _find_label_column(cells)
                if label_col is None:
                    continue
                current_col, prior_col = _find_value_columns(cells, label_col)

            label_text = _cell_text(cells.get(label_col))
            if not label_text:
                continue

            current_raw = _cell_text(cells.get(current_col)) if current_col is not None else None
            prior_raw = _cell_text(cells.get(prior_col)) if prior_col is not None else None

            rows.append({
                "label": label_text,
                "current": current_raw,
                "prior": prior_raw,
                "page": str(table.page_number),
                "is_header": False,
            })

    return rows


def _cell_text(cell: dict[str, Any] | None) -> str | None:
    if cell is None:
        return None
    raw_text = cell.get("raw_text")
    if raw_text is None:
        return None
    return str(raw_text).strip() or None


def _check_completeness(
    statement_rows: dict[str, list[dict[str, Any]]],
    *,
    market: str = "a_share",
) -> list[str]:
    errors: list[str] = []
    required_patterns = REQUIRED_LABEL_PATTERNS_BY_MARKET.get(
        market,
        REQUIRED_LABEL_PATTERNS_BY_MARKET["a_share"],
    )

    for role in REQUIRED_STATEMENT_ROLES:
        rows = statement_rows.get(role, [])
        if not rows:
            errors.append(f"缺少报表: {STATEMENT_SHEET_NAMES.get(role, role)}")
            continue

        all_labels = "\n".join(
            row["label"] or "" for row in rows if not row.get("is_header")
        )

        for pattern_group in required_patterns.get(role, []):
            if not any(
                _contains_pattern(all_labels, pattern) for pattern in pattern_group
            ):
                patterns_str = " / ".join(pattern_group)
                errors.append(
                    f"{STATEMENT_SHEET_NAMES.get(role, role)} 缺少关键科目: {patterns_str}"
                )

    return errors


def _apply_statement_formatting(ws, *, num_columns: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_columns)}1"

    for col_idx in range(1, num_columns + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    ws.column_dimensions["A"].width = 42

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row or 2):
        for cell in row:
            if cell.column > 1:
                cell.number_format = AMOUNT_FORMAT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = LABEL_ALIGNMENT


def _write_statement_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)

    for row in rows:
        if row.get("is_header"):
            continue
        current_val = _parse_decimal(row.get("current"))
        prior_val = _parse_decimal(row.get("prior"))
        ws.append([
            _excel_safe_text(row.get("label")),
            current_val if current_val is not None else _excel_safe_text(row.get("current")),
            prior_val if prior_val is not None else _excel_safe_text(row.get("prior")),
            _excel_safe_text(row.get("page")),
        ])

    _apply_statement_formatting(ws, num_columns=len(columns))


def _write_notes_sheet(
    wb: Workbook,
    *,
    metadata: dict[str, str],
    completeness_errors: list[str],
    table_sources: dict[str, list[dict[str, str]]],
) -> None:
    ws = wb.create_sheet("说明")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    ws.append(["项目", "内容"])
    ws[1][0].font = HEADER_FONT
    ws[1][1].font = HEADER_FONT

    ws.append(["来源PDF", metadata.get("pdf_path", "")])
    ws.append(["report_id", metadata.get("report_id", "")])
    ws.append(["extraction_run_id", metadata.get("extraction_run_id", "")])
    ws.append(["市场", metadata.get("market", "")])
    ws.append(["公司代码", metadata.get("company_id", "")])
    ws.append(["公司名称", metadata.get("company_name", "")])
    ws.append(["财年", metadata.get("fiscal_year", "")])
    ws.append(["报告类型", metadata.get("report_type", "")])
    ws.append(["数据来源", "原始抽取 (raw_tables / raw_cells)"])
    ws.append(["完整性检查", "通过" if not completeness_errors else "失败"])
    if completeness_errors:
        for error in completeness_errors:
            ws.append(["完整性错误", error])

    ws.append(["", ""])
    ws.append(["表格来源", ""])
    for role, sources in table_sources.items():
        sheet_name = STATEMENT_SHEET_NAMES.get(role, role)
        pages = ", ".join(
            sorted(set(str(src["page"]) for src in sources))
        )
        table_ids = ", ".join(src["table_id"] for src in sources)
        ws.append([f"{sheet_name} - 页码", pages])
        ws.append([f"{sheet_name} - 表格ID", table_ids])

    ws.freeze_panes = "A2"


def export_statement_workbook(
    conn: Connection,
    extraction_run_id: str,
    output_path: str | Path | None = None,
) -> Path:
    report = conn.execute(
        """
        select reports.report_id, reports.market, reports.company_id,
               reports.company_name, reports.fiscal_year, reports.report_type,
               reports.stored_pdf_path
        from extraction_runs
        join reports on reports.report_id = extraction_runs.report_id
        where extraction_runs.extraction_run_id = ?
        """,
        (extraction_run_id,),
    ).fetchone()
    if report is None:
        raise ValueError(f"Unknown extraction_run_id: {extraction_run_id}")

    report_id = str(report[0])
    market = str(report[1])
    company_id = str(report[2]) if report[2] else ""
    company_name = str(report[3]) if report[3] else ""
    fiscal_year = report[4]
    report_type = str(report[5]) if report[5] else ""
    pdf_path = str(report[6])

    source_tables = _source_tables(conn, extraction_run_id)
    groups = _statement_groups(source_tables)

    statement_rows: dict[str, list[dict[str, Any]]] = {}
    table_sources: dict[str, list[dict[str, str]]] = {}

    for group in groups:
        if group.table_role not in STATEMENT_METADATA:
            continue
        statement_type, _period_basis = STATEMENT_METADATA[group.table_role]
        rows = _reconstruct_rows(
            conn,
            group.tables,
            statement_type=statement_type,
        )
        # Merge across groups — the same statement may span multiple
        # page groups (e.g. parent-scope pages then consolidated-scope
        # pages).  The last header wins; data rows are appended.
        if group.table_role in statement_rows:
            existing = statement_rows[group.table_role]
            # Drop duplicate header rows from continuation groups.
            merged = existing[:]
            for row in rows:
                if row.get("is_header"):
                    continue
                merged.append(row)
            statement_rows[group.table_role] = merged
        else:
            statement_rows[group.table_role] = rows
        if group.table_role in table_sources:
            table_sources[group.table_role].extend(
                {"table_id": t.raw_table_id, "page": str(t.page_number)}
                for t in group.tables
            )
        else:
            table_sources[group.table_role] = [
                {"table_id": t.raw_table_id, "page": str(t.page_number)}
                for t in group.tables
            ]

    completeness_errors = _check_completeness(statement_rows, market=market)
    if completeness_errors:
        error_message = "\n".join(completeness_errors)
        raise ValueError(f"报表不完整，无法导出:\n{error_message}")

    if output_path is None:
        output_path = Path(
            f"data/review_exports/{company_id or 'unknown'}_"
            f"{fiscal_year or '0000'}_{report_type or 'unknown'}_"
            f"{extraction_run_id}_statements.xlsx"
        )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    bs_rows = statement_rows.get("statement.balance_sheet", [])
    _write_statement_sheet(wb, "资产负债表", BS_COLUMNS, bs_rows)

    is_rows = statement_rows.get("statement.income_statement", [])
    _write_statement_sheet(wb, "利润表", IS_CF_COLUMNS, is_rows)

    cf_rows = statement_rows.get("statement.cash_flow", [])
    _write_statement_sheet(wb, "现金流量表", IS_CF_COLUMNS, cf_rows)

    metadata = {
        "pdf_path": pdf_path,
        "report_id": report_id,
        "extraction_run_id": extraction_run_id,
        "market": market,
        "company_id": company_id,
        "company_name": company_name,
        "fiscal_year": str(fiscal_year) if fiscal_year else "",
        "report_type": report_type,
    }
    _write_notes_sheet(
        wb,
        metadata=metadata,
        completeness_errors=completeness_errors,
        table_sources=table_sources,
    )

    wb.save(output_path)
    return output_path
