from decimal import Decimal

from openpyxl import load_workbook

from fin_report_extractor.review_workbook import (
    CORRECTION_COLUMNS,
    export_review_workbook,
    read_corrections,
)


REQUIRED_METADATA = {
    "workbook_schema_version": "1",
    "report_id": "report-1",
    "extraction_run_id": "run-1",
    "review_export_id": "review-1",
    "rule_pack_version": "rules-1",
    "exported_at": "2026-05-31T00:00:00Z",
}


def test_export_review_workbook_creates_required_sheets(tmp_path):
    path = tmp_path / "review.xlsx"

    export_review_workbook(
        path,
        metadata=REQUIRED_METADATA,
        failures=[],
        raw_rows=[],
    )

    wb = load_workbook(path)

    assert wb.sheetnames == [
        "summary",
        "validation_failures",
        "balance_sheet_raw",
        "income_statement_raw",
        "cash_flow_raw",
        "notes_revenue_raw",
        "corrections",
        "_metadata",
    ]


def test_export_review_workbook_writes_review_columns(tmp_path):
    path = tmp_path / "review.xlsx"

    export_review_workbook(
        path,
        metadata=REQUIRED_METADATA,
        failures=[
            {
                "validation_result_id": "vr-1",
                "rule_id": "balance_sheet.assets_equal_liabilities_plus_equity",
                "severity": "error",
                "status": "failed",
                "message": "资产总计不平。",
                "involved_fact_ids": ["fact-1", "fact-2"],
                "lhs_value": Decimal("100"),
                "rhs_value": Decimal("90"),
                "difference_value": Decimal("10"),
                "source_pages": [12],
                "suggested_action": "在 corrections sheet 填写修正值",
            }
        ],
        raw_rows=[
            {
                "target_sheet": "balance_sheet_raw",
                "fact_id": "fact-1",
                "raw_table_id": "table-1",
                "raw_cell_id": "cell-1",
                "extractor_name": "fixture",
                "page_number": 12,
                "table_index_on_page": 1,
                "row_index": 2,
                "column_index": 3,
                "cell_bbox_json": "[0,0,1,1]",
                "table_role": "statement.balance_sheet",
                "statement_scope": "consolidated",
                "period_basis": "instant",
                "period_start": None,
                "period_end": None,
                "instant_date": "2025-12-31",
                "raw_label": "资产总计",
                "normalized_concept_id": "total_assets",
                "raw_value": "100",
                "raw_unit": "元",
                "currency": "CNY",
                "scale_factor": Decimal("1"),
                "normalized_value": Decimal("100"),
                "validation_status": "failed",
                "review_hint": "检查缺失科目",
                "display_col_1": "资产总计",
            }
        ],
    )

    wb = load_workbook(path)

    assert [cell.value for cell in wb["validation_failures"][1]] == [
        "validation_result_id",
        "rule_id",
        "severity",
        "status",
        "message",
        "involved_fact_ids",
        "lhs_value",
        "rhs_value",
        "difference_value",
        "source_pages",
        "suggested_action",
    ]
    assert wb["balance_sheet_raw"]["A2"].value == "fact-1"
    assert wb["income_statement_raw"].max_row == 1
    assert [cell.value for cell in wb["corrections"][1]] == CORRECTION_COLUMNS


def test_read_corrections_reads_user_rows(tmp_path):
    path = tmp_path / "review.xlsx"
    export_review_workbook(
        path,
        metadata=REQUIRED_METADATA,
        failures=[],
        raw_rows=[],
    )
    wb = load_workbook(path)
    ws = wb["corrections"]
    ws.append(
        [
            "fact-1",
            "correct",
            "101.25",
            "元",
            None,
            None,
            None,
            None,
            "人工修正数值",
        ]
    )
    ws.append(
        [
            "fact-2",
            "remap",
            None,
            None,
            "total_liabilities",
            "instant",
            "consolidated",
            "statement.balance_sheet",
            "人工修正科目",
        ]
    )
    wb.save(path)

    rows = read_corrections(path)

    assert rows == [
        {
            "fact_id": "fact-1",
            "correction_action": "correct",
            "corrected_value": "101.25",
            "corrected_unit": "元",
            "normalized_concept_id": None,
            "period_basis": None,
            "statement_scope": None,
            "table_role": None,
            "correction_reason": "人工修正数值",
        },
        {
            "fact_id": "fact-2",
            "correction_action": "remap",
            "corrected_value": None,
            "corrected_unit": None,
            "normalized_concept_id": "total_liabilities",
            "period_basis": "instant",
            "statement_scope": "consolidated",
            "table_role": "statement.balance_sheet",
            "correction_reason": "人工修正科目",
        },
    ]


def test_read_corrections_ignores_raw_sheet_user_edits(tmp_path):
    path = tmp_path / "review.xlsx"
    export_review_workbook(
        path,
        metadata=REQUIRED_METADATA,
        failures=[],
        raw_rows=[
            {
                "target_sheet": "balance_sheet_raw",
                "fact_id": "fact-original",
                "raw_label": "资产总计",
                "raw_value": "100",
            }
        ],
    )
    wb = load_workbook(path)
    wb["balance_sheet_raw"]["A2"] = "fact-from-raw-user-edit"
    wb["corrections"].append(
        [
            "fact-from-corrections",
            "confirm",
            None,
            None,
            None,
            None,
            None,
            None,
            "人工确认",
        ]
    )
    wb.save(path)

    rows = read_corrections(path)

    assert rows[0]["fact_id"] == "fact-from-corrections"
    assert rows[0]["correction_action"] == "confirm"
