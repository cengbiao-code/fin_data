from pathlib import Path

import pytest

from fin_report_extractor.audit_db import connect_audit_db, initialize_audit_db
from fin_report_extractor.extraction_runs import create_extraction_run, persist_raw_tables
from fin_report_extractor.extractors import ExtractedCell, ExtractedTable
from fin_report_extractor.import_pdf import register_pdf
from fin_report_extractor.statement_workbook import (
    _adjust_columns_from_data,
    _check_completeness,
    _excel_safe_text,
    _is_numeric_text,
    _parse_decimal,
    export_statement_workbook,
)
from fin_report_extractor.table_classifier import classify_tables_for_run


def _insert_page(conn, report_id, page_number, text):
    conn.execute(
        """
        insert into pdf_pages (
          page_id, report_id, page_number, width, height, text_char_count,
          text_density, has_statement_keywords, page_text_sample
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            f"page-{page_number}",
            report_id,
            page_number,
            595.0,
            842.0,
            len(text),
            0.0002,
            1,
            text,
        ),
    )


def _a_share_data_row(row_index, label, current, prior=None, page_number=1):
    return [
        ExtractedCell(row_index, 0, None, None, page_number),
        ExtractedCell(row_index, 1, label, None, page_number),
        ExtractedCell(row_index, 2, None, None, page_number),
        ExtractedCell(row_index, 3, current, None, page_number),
        ExtractedCell(row_index, 4, None, None, page_number),
        ExtractedCell(row_index, 5, None, None, page_number),
        ExtractedCell(row_index, 6, prior, None, page_number),
        ExtractedCell(row_index, 7, None, None, page_number),
        ExtractedCell(row_index, 8, None, None, page_number),
    ]


def _a_share_narrow_data_row(row_index, label, current, prior=None, page_number=1):
    """5-column data row (0-4), simulating continuation pages with narrower tables."""
    return [
        ExtractedCell(row_index, 0, None, None, page_number),
        ExtractedCell(row_index, 1, label, None, page_number),
        ExtractedCell(row_index, 2, None, None, page_number),
        ExtractedCell(row_index, 3, current, None, page_number),
        ExtractedCell(row_index, 4, prior, None, page_number),
    ]


def _bs_header(page=1):
    return [
        ExtractedCell(0, 1, "项目", None, page),
        ExtractedCell(0, 3, "期末余额", None, page),
        ExtractedCell(0, 6, "期初余额", None, page),
    ]


def _is_cf_header(page=1):
    return [
        ExtractedCell(0, 1, "项目", None, page),
        ExtractedCell(0, 3, "本期发生额", None, page),
        ExtractedCell(0, 6, "上期发生额", None, page),
    ]


def _bs_header_offset(page=1):
    """A-share merged-cell header: title text shifted +1 col from data."""
    return [
        ExtractedCell(0, 1, "项目", None, page),
        ExtractedCell(0, 4, "期末余额", None, page),
        ExtractedCell(0, 7, "期初余额", None, page),
    ]


def _is_cf_header_offset(page=1):
    """A-share merged-cell header: title text shifted +1 col from data."""
    return [
        ExtractedCell(0, 1, "项目", None, page),
        ExtractedCell(0, 4, "本期发生额", None, page),
        ExtractedCell(0, 7, "上期发生额", None, page),
    ]


# ── parse_decimal / is_numeric_text ──────────────────────────────────────────


class TestParseDecimal:
    def test_parse_simple_integer(self):
        assert _parse_decimal("100") == 100.0

    def test_parse_with_commas(self):
        assert _parse_decimal("41,506,860,074.79") == 41506860074.79

    def test_parse_blank_is_none(self):
        assert _parse_decimal("") is None

    def test_parse_dash_is_none(self):
        assert _parse_decimal("--") is None

    def test_parse_parentheses_negative(self):
        assert _parse_decimal("(100.50)") == -100.50

    def test_parse_none_is_none(self):
        assert _parse_decimal(None) is None

    def test_is_numeric_text_identifies_number(self):
        assert _is_numeric_text("100") is True

    def test_is_numeric_text_rejects_label(self):
        assert _is_numeric_text("资产总计") is False

    def test_is_numeric_text_rejects_empty(self):
        assert _is_numeric_text("") is False

    def test_excel_safe_text_removes_illegal_control_characters(self):
        assert _excel_safe_text("abc\x01def\x0bg") == "abcdefg"


# ── adjust_columns_from_data ──────────────────────────────────────────────────


class TestAdjustColumnsFromData:
    def test_current_col_has_numeric_returns_verified(self):
        cells = {
            1: {"raw_text": "货币资金"},
            3: {"raw_text": "124,472,018,105.34"},
            6: {"raw_text": "113,900,461,797.94"},
        }
        current, prior, verified = _adjust_columns_from_data(cells, 1, 3, 6)
        assert current == 3
        assert prior == 6
        assert verified is True

    def test_offset_minus_one_corrects_merged_header_columns(self):
        # Real A-share PDF scenario: header text at col 4/7 (due to merged
        # cells), but actual data values sit at col 3/6.
        cells = {
            1: {"raw_text": "营业总收入"},
            3: {"raw_text": "41,639,050,426.79"},
            6: {"raw_text": "36,595,551,580.44"},
        }
        current, prior, verified = _adjust_columns_from_data(cells, 1, 4, 7)
        assert current == 3, "current should shift left from 4 to 3"
        assert prior == 6, "prior should shift left from 7 to 6"
        assert verified is True

    def test_section_header_row_returns_not_verified(self):
        # Section headers like "流动资产：" have no numeric cells —
        # verification should report False so callers know to keep looking.
        cells = {
            1: {"raw_text": "流动资产："},
            3: {"raw_text": None},
            6: {"raw_text": None},
        }
        current, prior, verified = _adjust_columns_from_data(cells, 1, 4, 7)
        assert current == 4  # unchanged
        assert prior == 7   # unchanged
        assert verified is False

    def test_data_at_original_columns_keeps_them(self):
        cells = {
            1: {"raw_text": "货币资金"},
            3: {"raw_text": "124,472,018,105.34"},
            6: {"raw_text": "113,900,461,797.94"},
        }
        current, prior, verified = _adjust_columns_from_data(cells, 1, 3, 6)
        assert current == 3
        assert prior == 6
        assert verified is True

    def test_none_current_col_returns_unchanged(self):
        cells = {
            1: {"raw_text": "货币资金"},
            3: {"raw_text": "100.00"},
        }
        current, prior, verified = _adjust_columns_from_data(cells, 1, None, 6)
        assert current is None
        assert prior == 6
        assert verified is False


# ── completeness checks ──────────────────────────────────────────────────────


def _minimal_is_cf():
    return {
        "statement.income_statement": [
            {"label": "净利润", "is_header": False},
            {"label": "综合收益总额", "is_header": False},
            {"label": "基本每股收益", "is_header": False},
        ],
        "statement.cash_flow": [
            {"label": "经营活动产生的现金流量净额", "is_header": False},
            {"label": "投资活动产生的现金流量净额", "is_header": False},
            {"label": "筹资活动产生的现金流量净额", "is_header": False},
            {"label": "期末现金及现金等价物余额", "is_header": False},
        ],
    }


class TestCompletenessCheck:
    def test_all_present_passes(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "资产总计", "is_header": False},
                {"label": "负债合计", "is_header": False},
                {"label": "所有者权益合计", "is_header": False},
            ],
            **_minimal_is_cf(),
        }
        assert _check_completeness(rows) == []

    def test_missing_statement_role_reports_error(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "资产总计", "is_header": False},
                {"label": "负债合计", "is_header": False},
                {"label": "所有者权益合计", "is_header": False},
            ],
        }
        errors = _check_completeness(rows)
        assert any("利润表" in e for e in errors)
        assert any("现金流量表" in e for e in errors)

    def test_missing_key_label_reports_error(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "some other label", "is_header": False},
            ],
            **_minimal_is_cf(),
        }
        errors = _check_completeness(rows)
        assert any("资产总计" in e for e in errors)

    def test_alternative_label_patterns_accepted(self):
        rows_equity = {
            "statement.balance_sheet": [
                {"label": "资产总计", "is_header": False},
                {"label": "负债合计", "is_header": False},
                {"label": "股东权益合计", "is_header": False},
            ],
        }
        assert _check_completeness(rows_equity | _minimal_is_cf()) == []

        rows_sum = {
            "statement.balance_sheet": [
                {"label": "资产总计", "is_header": False},
                {"label": "负债合计", "is_header": False},
                {"label": "负债和所有者权益总计", "is_header": False},
            ],
        }
        assert _check_completeness(rows_sum | _minimal_is_cf()) == []

    def test_empty_rows_reports_missing_statement(self):
        rows = {
            "statement.balance_sheet": [],
            **_minimal_is_cf(),
        }
        errors = _check_completeness(rows)
        assert any("资产负债表" in e for e in errors)

    def test_hk_statement_labels_are_accepted(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "資產總額", "is_header": False},
                {"label": "負債總額", "is_header": False},
                {"label": "權益總額", "is_header": False},
            ],
            "statement.income_statement": [
                {"label": "期內盈利", "is_header": False},
                {"label": "本公司權益持有人應佔每股盈利 －基本", "is_header": False},
                {"label": "經營盈利", "is_header": False},
            ],
            "statement.cash_flow": [
                {"label": "經營活動所得現金流量淨額", "is_header": False},
                {"label": "投資活動耗用現金流量淨額", "is_header": False},
                {"label": "融資活動（耗用）╱所得現金流量淨額", "is_header": False},
                {"label": "期末的現金及現金等價物", "is_header": False},
            ],
        }

        assert _check_completeness(rows, market="hk") == []

    def test_hk_loss_statement_labels_are_accepted(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "資產總額", "is_header": False},
                {"label": "負債總額", "is_header": False},
                {"label": "權益總額", "is_header": False},
            ],
            "statement.income_statement": [
                {"label": "年內（虧損）╱", "is_header": False},
                {"label": "每股基本（虧損）╱", "is_header": False},
                {"label": "經營（虧損）", "is_header": False},
            ],
            "statement.cash_flow": [
                {"label": "經營活動（所用）╱ 所得現金流量淨額", "is_header": False},
                {"label": "投資活動所得現金流量淨額", "is_header": False},
                {"label": "融資活動所得╱（所用）現金流量淨額", "is_header": False},
                {"label": "年末現金及現金等價物", "is_header": False},
            ],
        }

        assert _check_completeness(rows, market="hk") == []

    def test_hk_exact_label_match_required_for_total_liabilities(self):
        """權益及負債總額 should NOT satisfy 負債總額 or 權益總額 via substring."""
        rows = {
            "statement.balance_sheet": [
                {"label": "資產總額", "is_header": False},
                {"label": "權益及負債總額", "is_header": False},
            ],
            **_minimal_is_cf(),
        }
        errors = _check_completeness(rows, market="hk")
        assert any("負債總額" in e for e in errors)
        assert any("權益總額" in e for e in errors)

    def test_us_statement_labels_are_accepted(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "Total assets", "is_header": False},
                {"label": "Total liabilities", "is_header": False},
                {"label": "Total stockholders' equity", "is_header": False},
            ],
            "statement.income_statement": [
                {"label": "Net income", "is_header": False},
                {"label": "Earnings per share", "is_header": False},
                {"label": "Operating income", "is_header": False},
            ],
            "statement.cash_flow": [
                {"label": "Net cash provided by operating activities", "is_header": False},
                {"label": "Net cash used in investing activities", "is_header": False},
                {"label": "Net cash provided by financing activities", "is_header": False},
                {"label": "Cash and cash equivalents, end of period", "is_header": False},
            ],
        }
        assert _check_completeness(rows, market="us") == []

    def test_us_statement_equity_variant_accepted(self):
        rows = {
            "statement.balance_sheet": [
                {"label": "Total assets", "is_header": False},
                {"label": "Total liabilities", "is_header": False},
                {"label": "Total shareholders' equity", "is_header": False},
            ],
            "statement.income_statement": [
                {"label": "Net earnings", "is_header": False},
                {"label": "EPS", "is_header": False},
                {"label": "Income from operations", "is_header": False},
            ],
            "statement.cash_flow": [
                {"label": "Net cash from operating activities", "is_header": False},
                {"label": "Net cash used in investing activities", "is_header": False},
                {"label": "Net cash from financing activities", "is_header": False},
                {"label": "Cash and cash equivalents", "is_header": False},
            ],
        }
        assert _check_completeness(rows, market="us") == []


# ── helper: create full three-statement DB ───────────────────────────────────


def _setup_three_statement_data(tmp_path, *, market="a_share"):
    conn = connect_audit_db(tmp_path / "audit.sqlite")
    initialize_audit_db(conn)
    pdf = tmp_path / "sample.pdf"
    pdf.write_bytes(b"%PDF-1.4 sample\n")
    report_id = register_pdf(
        conn, pdf,
        stored_pdf_path=str(pdf),
        market=market,
        company_id="000651",
        company_name="格力电器",
        fiscal_year=2025,
        report_type="quarterly",
    )
    _insert_page(conn, report_id, 1, "合并资产负债表\n单位：元")
    _insert_page(conn, report_id, 2, "合并利润表\n单位：元")
    _insert_page(conn, report_id, 3, "合并现金流量表\n单位：元")
    run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

    bs_rows = _bs_header(1)[:]
    bs_rows.extend(_a_share_data_row(1, "资产总计", "394,568,798,089.75", "368,031,704,522.86", 1))
    bs_rows.extend(_a_share_data_row(2, "负债合计", "247,276,037,775.29", "226,518,009,574.89", 1))
    bs_rows.extend(_a_share_data_row(3, "所有者权益合计", "147,292,760,314.46", "141,513,694,947.97", 1))

    is_rows = _is_cf_header(2)[:]
    is_rows.extend(_a_share_data_row(1, "其中：营业收入", "41,506,860,074.79", "36,335,114,366.16", 2))
    is_rows.extend(_a_share_data_row(2, "五、净利润", "5,941,618,723.39", "4,664,908,787.11", 2))
    is_rows.extend(_a_share_data_row(3, "综合收益总额", "5,921,990,942.03", "4,651,233,646.19", 2))
    is_rows.extend(_a_share_data_row(4, "基本每股收益", "1.37", "1.08", 2))

    cf_rows = _is_cf_header(3)[:]
    cf_rows.extend(_a_share_data_row(1, "经营活动产生的现金流量净额", "11,001,218,583.01", None, 3))
    cf_rows.extend(_a_share_data_row(2, "投资活动产生的现金流量净额", "-5,234,567,890.12", None, 3))
    cf_rows.extend(_a_share_data_row(3, "筹资活动产生的现金流量净额", "-2,100,000,000.00", None, 3))
    cf_rows.extend(_a_share_data_row(4, "期末现金及现金等价物余额", "50,000,000,000.00", None, 3))

    persist_raw_tables(
        conn, report_id, run_id,
        [
            ExtractedTable("pdfplumber", 1, 0, None, bs_rows, {}),
            ExtractedTable("pdfplumber", 2, 0, None, is_rows, {}),
            ExtractedTable("pdfplumber", 3, 0, None, cf_rows, {}),
        ],
    )
    classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
    return conn, report_id, run_id


# ── full workbook export ─────────────────────────────────────────────────────


class TestExportStatementWorkbook:
    def test_workbook_has_expected_sheets(self, tmp_path):
        conn, _report_id, run_id = _setup_three_statement_data(tmp_path)
        output = tmp_path / "output.xlsx"
        try:
            path = export_statement_workbook(conn, run_id, output_path=output)
            assert path == output
            assert output.exists()

            from openpyxl import load_workbook
            wb = load_workbook(output)
            assert "资产负债表" in wb.sheetnames
            assert "利润表" in wb.sheetnames
            assert "现金流量表" in wb.sheetnames
            assert "说明" in wb.sheetnames
        finally:
            conn.close()

    def test_balance_sheet_has_row_data(self, tmp_path):
        conn, _report_id, run_id = _setup_three_statement_data(tmp_path)
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["资产负债表"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            assert rows[0] == ("项目", "期末余额", "期初余额", "来源页")
            labels = [r[0] for r in rows[1:]]
            assert "资产总计" in labels
            assert "负债合计" in labels
            assert "所有者权益合计" in labels
        finally:
            conn.close()

    def test_amounts_are_numeric(self, tmp_path):
        conn, _report_id, run_id = _setup_three_statement_data(tmp_path)
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["利润表"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            revenue_row = [r for r in rows if r[0] and "营业收入" in str(r[0])][0]
            assert isinstance(revenue_row[1], (int, float)), (
                f"Expected numeric, got {type(revenue_row[1])}"
            )
        finally:
            conn.close()

    def test_header_row_is_frozen(self, tmp_path):
        conn, _report_id, run_id = _setup_three_statement_data(tmp_path)
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            assert wb["资产负债表"].freeze_panes == "A2"
        finally:
            conn.close()

    def test_notes_sheet_has_metadata(self, tmp_path):
        conn, _report_id, run_id = _setup_three_statement_data(tmp_path)
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["说明"]
            rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
            all_text = " ".join(str(c) for r in rows for c in r if c)
            assert "格力电器" in all_text
            assert "000651" in all_text
            assert "原始抽取" in all_text
        finally:
            conn.close()

    def test_missing_statement_raises_value_error(self, tmp_path):
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="a_share", company_id="TEST", fiscal_year=2025, report_type="annual",
        )
        _insert_page(conn, report_id, 1, "合并资产负债表\n单位：元")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})
        bs_rows = _bs_header(1)[:]
        bs_rows.extend(_a_share_data_row(1, "资产总计", "100", "90", 1))
        bs_rows.extend(_a_share_data_row(2, "负债合计", "40", "30", 1))
        bs_rows.extend(_a_share_data_row(3, "所有者权益合计", "60", "60", 1))
        persist_raw_tables(
            conn, report_id, run_id,
            [ExtractedTable("pdfplumber", 1, 0, None, bs_rows, {})],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            with pytest.raises(ValueError, match="报表不完整"):
                export_statement_workbook(conn, run_id, output_path=output)
            assert not output.exists()
        finally:
            conn.close()

    def test_missing_key_label_raises_value_error(self, tmp_path):
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="a_share", company_id="TEST", fiscal_year=2025, report_type="annual",
        )
        _insert_page(conn, report_id, 1, "合并资产负债表\n单位：元")
        _insert_page(conn, report_id, 2, "合并利润表\n单位：元")
        _insert_page(conn, report_id, 3, "合并现金流量表\n单位：元")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

        bs_rows = _bs_header(1)[:]
        bs_rows.extend(_a_share_data_row(1, "无形资产", "100", "90", 1))
        bs_rows.extend(_a_share_data_row(2, "短期借款", "40", "30", 1))

        is_rows = _is_cf_header(2)[:]
        is_rows.extend(_a_share_data_row(1, "净利润", "100", "90", 2))
        is_rows.extend(_a_share_data_row(2, "综合收益总额", "100", "90", 2))
        is_rows.extend(_a_share_data_row(3, "基本每股收益", "1.00", "0.90", 2))

        cf_rows = _is_cf_header(3)[:]
        cf_rows.extend(_a_share_data_row(1, "经营活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(2, "投资活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(3, "筹资活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(4, "期末现金及现金等价物余额", "100", None, 3))

        persist_raw_tables(
            conn, report_id, run_id,
            [
                ExtractedTable("pdfplumber", 1, 0, None, bs_rows, {}),
                ExtractedTable("pdfplumber", 2, 0, None, is_rows, {}),
                ExtractedTable("pdfplumber", 3, 0, None, cf_rows, {}),
            ],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            with pytest.raises(ValueError, match="资产总计"):
                export_statement_workbook(conn, run_id, output_path=output)
            assert not output.exists()
        finally:
            conn.close()

    def test_cross_page_continuation_included(self, tmp_path):
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="a_share", company_id="000651", company_name="格力电器",
            fiscal_year=2025, report_type="quarterly",
        )
        _insert_page(conn, report_id, 1, "合并资产负债表\n单位：元")
        _insert_page(conn, report_id, 2, "合并利润表\n单位：元")
        _insert_page(conn, report_id, 3, "合并现金流量表\n单位：元")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

        bs_page1 = _bs_header(1)[:]
        bs_page1.extend(_a_share_data_row(1, "资产总计", "100", "90", 1))
        bs_page1.extend(_a_share_data_row(2, "负债合计", "40", "30", 1))
        bs_page2 = _a_share_data_row(0, "所有者权益合计", "60", "60", 2)

        is_rows = _is_cf_header(2)[:]
        is_rows.extend(_a_share_data_row(1, "净利润", "100", "90", 2))
        is_rows.extend(_a_share_data_row(2, "综合收益总额", "100", "90", 2))
        is_rows.extend(_a_share_data_row(3, "基本每股收益", "1.00", "0.90", 2))

        cf_rows = _is_cf_header(3)[:]
        cf_rows.extend(_a_share_data_row(1, "经营活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(2, "投资活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(3, "筹资活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(4, "期末现金及现金等价物余额", "100", None, 3))

        persist_raw_tables(
            conn, report_id, run_id,
            [
                ExtractedTable("pdfplumber", 1, 0, None, bs_page1, {}),
                ExtractedTable("pdfplumber", 2, 0, None, bs_page2, {}),
                ExtractedTable("pdfplumber", 2, 1, None, is_rows, {}),
                ExtractedTable("pdfplumber", 3, 0, None, cf_rows, {}),
            ],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            path = export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["资产负债表"]
            labels = [
                str(r[0]) for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] is not None
            ]
            assert "资产总计" in labels
            assert "负债合计" in labels
            assert "所有者权益合计" in labels
        finally:
            conn.close()

    def test_offset_header_columns_extracts_data(self, tmp_path):
        """A-share merged cells often shift header text +1 col from data."""
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="a_share", company_id="000651", company_name="格力电器",
            fiscal_year=2025, report_type="quarterly",
        )
        _insert_page(conn, report_id, 1, "合并资产负债表\n单位：元")
        _insert_page(conn, report_id, 2, "合并利润表\n单位：元")
        _insert_page(conn, report_id, 3, "合并现金流量表\n单位：元")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

        # Header at col 4/7 but data at col 3/6 (real A-share pattern).
        bs_rows = _bs_header_offset(1)[:]
        bs_rows.extend(_a_share_data_row(1, "资产总计", "100", "90", 1))
        bs_rows.extend(_a_share_data_row(2, "负债合计", "40", "30", 1))
        bs_rows.extend(_a_share_data_row(3, "所有者权益合计", "60", "60", 1))

        is_rows = _is_cf_header_offset(2)[:]
        is_rows.extend(_a_share_data_row(1, "净利润", "50", "40", 2))
        is_rows.extend(_a_share_data_row(2, "综合收益总额", "50", "40", 2))
        is_rows.extend(_a_share_data_row(3, "基本每股收益", "1.00", "0.90", 2))

        cf_rows = _is_cf_header_offset(3)[:]
        cf_rows.extend(_a_share_data_row(1, "经营活动产生的现金流量净额", "30", None, 3))
        cf_rows.extend(_a_share_data_row(2, "投资活动产生的现金流量净额", "20", None, 3))
        cf_rows.extend(_a_share_data_row(3, "筹资活动产生的现金流量净额", "10", None, 3))
        cf_rows.extend(_a_share_data_row(4, "期末现金及现金等价物余额", "60", None, 3))

        persist_raw_tables(
            conn, report_id, run_id,
            [
                ExtractedTable("pdfplumber", 1, 0, None, bs_rows, {}),
                ExtractedTable("pdfplumber", 2, 0, None, is_rows, {}),
                ExtractedTable("pdfplumber", 3, 0, None, cf_rows, {}),
            ],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["资产负债表"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            asset_row = [r for r in rows if r[0] and "资产总计" in str(r[0])][0]
            assert isinstance(asset_row[1], (int, float)), (
                f"Expected numeric, got {type(asset_row[1])}"
            )
            assert asset_row[1] == 100
            assert asset_row[2] == 90
        finally:
            conn.close()

    def test_blank_values_preserved_as_blank(self, tmp_path):
        conn, _report_id, run_id = _setup_three_statement_data(tmp_path)
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)
            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["现金流量表"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            ocf_row = [r for r in rows if r[0] and "经营活动" in str(r[0])][0]
            # prior period value for CF rows is None → should be blank
            assert ocf_row[2] is None or ocf_row[2] == ""
        finally:
            conn.close()

    def test_cross_page_cashflow_narrow_continuation(self, tmp_path):
        """CF continuation page with fewer columns must still extract prior values."""
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="a_share", company_id="000651", company_name="格力电器",
            fiscal_year=2025, report_type="quarterly",
        )
        _insert_page(conn, report_id, 6, "合并资产负债表\n单位：元")
        _insert_page(conn, report_id, 7, "合并利润表\n单位：元")
        _insert_page(conn, report_id, 8, "合并现金流量表\n单位：元")
        _insert_page(conn, report_id, 9, "母公司现金流量表\n单位：元")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

        # Minimal BS and IS to satisfy completeness check.
        bs_rows = _is_cf_header(6)[:]
        bs_rows.extend(_a_share_data_row(1, "资产总计", "100", "90", 6))
        bs_rows.extend(_a_share_data_row(2, "负债合计", "40", "30", 6))
        bs_rows.extend(_a_share_data_row(3, "所有者权益合计", "60", "60", 6))

        is_rows = _is_cf_header(7)[:]
        is_rows.extend(_a_share_data_row(1, "净利润", "50", "40", 7))
        is_rows.extend(_a_share_data_row(2, "综合收益总额", "50", "40", 7))
        is_rows.extend(_a_share_data_row(3, "基本每股收益", "1.00", "0.90", 7))

        # Page 8: consolidated CF, 9 columns (0-8), current at col 3, prior at col 6.
        cf_page8 = _is_cf_header(8)[:]
        cf_page8.extend(_a_share_data_row(1, "经营活动产生的现金流量净额", "100", "80", 8))
        cf_page8.extend(_a_share_data_row(2, "投资活动产生的现金流量净额", "50", "60", 8))

        # Page 9: parent CF continuation, only 5 columns (0-4),
        # prior values sit at col 4 (rightmost available column).
        cf_page9 = _a_share_narrow_data_row(0, "筹资活动产生的现金流量净额", "30", "20", 9)
        cf_page9.extend(_a_share_narrow_data_row(1, "期末现金及现金等价物余额", "200", "180", 9))

        persist_raw_tables(
            conn, report_id, run_id,
            [
                ExtractedTable("pdfplumber", 6, 0, None, bs_rows, {}),
                ExtractedTable("pdfplumber", 7, 0, None, is_rows, {}),
                ExtractedTable("pdfplumber", 8, 0, None, cf_page8, {}),
                ExtractedTable("pdfplumber", 9, 0, None, cf_page9, {}),
            ],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["现金流量表"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            # Rows from page 8 should have both current and prior.
            ocf_row = [r for r in rows if r[0] and "经营活动" in str(r[0])][0]
            assert ocf_row[1] == 100
            assert ocf_row[2] == 80

            # Rows from page 9 should also have both current and prior.
            financing_row = [r for r in rows if r[0] and "筹资活动" in str(r[0])][0]
            assert financing_row[1] == 30
            assert financing_row[2] == 20

            cash_row = [r for r in rows if r[0] and "期末现金" in str(r[0])][0]
            assert cash_row[1] == 200
            assert cash_row[2] == 180
        finally:
            conn.close()

    def test_segment_note_after_cashflow_is_not_appended(self, tmp_path):
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="hk", company_id="0700.HK", company_name="Tencent",
            fiscal_year=2025, report_type="annual",
        )
        _insert_page(conn, report_id, 1, "簡明綜合收益表\n人民幣百萬元")
        _insert_page(conn, report_id, 2, "簡明綜合財務狀況表\n人民幣百萬元")
        _insert_page(conn, report_id, 3, "簡明綜合現金流量表\n人民幣百萬元")
        _insert_page(conn, report_id, 4, "分部資料及收入\n與簡明綜合收益表採用一致的方式計量")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pymupdf": "test"})

        is_rows = _is_cf_header(1)[:]
        is_rows.extend(_a_share_data_row(1, "期內盈利", "100", "90", 1))
        is_rows.extend(_a_share_data_row(2, "每股盈利", "1.00", "0.90", 1))
        is_rows.extend(_a_share_data_row(3, "經營盈利", "50", "45", 1))

        bs_rows = _is_cf_header(2)[:]
        bs_rows.extend(_a_share_data_row(1, "資產總額", "500", "450", 2))
        bs_rows.extend(_a_share_data_row(2, "負債總額", "200", "180", 2))
        bs_rows.extend(_a_share_data_row(3, "權益總額", "300", "270", 2))

        cf_rows = _is_cf_header(3)[:]
        cf_rows.extend(_a_share_data_row(1, "經營活動所得現金流量淨額", "100", "80", 3))
        cf_rows.extend(_a_share_data_row(2, "投資活動耗用現金流量淨額", "-50", "-60", 3))
        cf_rows.extend(_a_share_data_row(3, "融資活動（耗用）╱所得現金流量淨額", "-30", "-20", 3))
        cf_rows.extend(_a_share_data_row(4, "期末的現金及現金等價物", "200", "180", 3))

        note_rows = [
            ExtractedCell(0, 0, "2", None, 4),
            ExtractedCell(0, 1, "分部資料及收入", None, 4),
            ExtractedCell(1, 0, "與簡明綜合收益表採用一致的方式計量", None, 4),
            ExtractedCell(2, 0, "分部收入", None, 4),
            ExtractedCell(2, 1, "196,458", None, 4),
        ]

        persist_raw_tables(
            conn, report_id, run_id,
            [
                ExtractedTable("pymupdf", 1, 0, None, is_rows, {}),
                ExtractedTable("pymupdf", 2, 0, None, bs_rows, {}),
                ExtractedTable("pymupdf", 3, 0, None, cf_rows, {}),
                ExtractedTable("pymupdf", 4, 0, None, note_rows, {}),
            ],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)

            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["现金流量表"]
            labels = [
                str(r[0]) for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] is not None
            ]
            assert "期末的現金及現金等價物" in labels
            assert "分部資料及收入" not in labels
            assert "分部收入" not in labels
        finally:
            conn.close()

    def test_non_simple_values_kept_as_text(self, tmp_path):
        conn = connect_audit_db(tmp_path / "audit.sqlite")
        initialize_audit_db(conn)
        pdf = tmp_path / "sample.pdf"
        pdf.write_bytes(b"%PDF-1.4 sample\n")
        report_id = register_pdf(
            conn, pdf, stored_pdf_path=str(pdf),
            market="a_share", company_id="TEST", fiscal_year=2025, report_type="annual",
        )
        _insert_page(conn, report_id, 1, "合并资产负债表\n单位：元")
        _insert_page(conn, report_id, 2, "合并利润表\n单位：元")
        _insert_page(conn, report_id, 3, "合并现金流量表\n单位：元")
        run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

        bs_rows = _bs_header(1)[:]
        bs_rows.extend(_a_share_data_row(1, "资产总计", "~100 million", "90", 1))
        bs_rows.extend(_a_share_data_row(2, "负债合计", "40", "30", 1))
        bs_rows.extend(_a_share_data_row(3, "所有者权益合计", "60", "60", 1))

        is_rows = _is_cf_header(2)[:]
        is_rows.extend(_a_share_data_row(1, "净利润", "100", "90", 2))
        is_rows.extend(_a_share_data_row(2, "综合收益总额", "100", "90", 2))
        is_rows.extend(_a_share_data_row(3, "基本每股收益", "1.00", "0.90", 2))

        cf_rows = _is_cf_header(3)[:]
        cf_rows.extend(_a_share_data_row(1, "经营活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(2, "投资活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(3, "筹资活动产生的现金流量净额", "100", None, 3))
        cf_rows.extend(_a_share_data_row(4, "期末现金及现金等价物余额", "100", None, 3))

        persist_raw_tables(
            conn, report_id, run_id,
            [
                ExtractedTable("pdfplumber", 1, 0, None, bs_rows, {}),
                ExtractedTable("pdfplumber", 2, 0, None, is_rows, {}),
                ExtractedTable("pdfplumber", 3, 0, None, cf_rows, {}),
            ],
        )
        classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
        output = tmp_path / "output.xlsx"
        try:
            export_statement_workbook(conn, run_id, output_path=output)
            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb["资产负债表"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            asset_row = [r for r in rows if r[0] and "资产总计" in str(r[0])][0]
            assert isinstance(asset_row[1], str), f"Expected text, got {type(asset_row[1])}"
        finally:
            conn.close()
