from pathlib import Path
import sqlite3

from fin_report_extractor.extractors import ExtractedCell, ExtractedTable
from fin_report_extractor.cli import main


def test_init_db_creates_audit_and_analytics_databases(tmp_path, capsys):
    audit_db = tmp_path / "nested" / "audit.sqlite"
    analytics_db = tmp_path / "nested" / "analytics.duckdb"

    main(
        [
            "init-db",
            "--audit-db",
            str(audit_db),
            "--analytics-db",
            str(analytics_db),
        ]
    )

    output = capsys.readouterr().out
    assert str(audit_db) in output
    assert str(analytics_db) in output

    sqlite_conn = sqlite3.connect(audit_db)
    try:
        audit_tables = {
            row[0]
            for row in sqlite_conn.execute(
                "select name from sqlite_master where type = 'table'"
            ).fetchall()
        }
    finally:
        sqlite_conn.close()

    assert "reports" in audit_tables
    assert analytics_db.exists()


def test_import_pdf_registers_report_with_metadata(tmp_path, capsys):
    audit_db = tmp_path / "db" / "audit.sqlite"
    pdf = tmp_path / "annual-report.pdf"
    pdf.write_bytes(b"%PDF-1.4 annual report\n")

    main(
        [
            "import-pdf",
            str(pdf),
            "--audit-db",
            str(audit_db),
            "--stored-pdf-path",
            "data/raw/annual-report.pdf",
            "--market",
            "a_share",
            "--company-id",
            "000001",
            "--company-name",
            "Example Corp",
            "--fiscal-year",
            "2025",
            "--report-type",
            "annual",
        ]
    )

    report_id = capsys.readouterr().out.strip()
    assert report_id

    conn = sqlite3.connect(audit_db)
    try:
        row = conn.execute(
            """
            select report_id, stored_pdf_path, market, company_id, company_name,
                   fiscal_year, report_type
            from reports
            """
        ).fetchone()
    finally:
        conn.close()

    assert row == (
        report_id,
        "data/raw/annual-report.pdf",
        "a_share",
        "000001",
        "Example Corp",
        2025,
        "annual",
    )


def test_extract_tables_persists_raw_tables_for_registered_report(
    tmp_path, capsys, monkeypatch
):
    audit_db = tmp_path / "db" / "audit.sqlite"
    pdf = tmp_path / "annual-report.pdf"
    pdf.write_bytes(b"%PDF-1.4 annual report\n")

    main(
        [
            "import-pdf",
            str(pdf),
            "--audit-db",
            str(audit_db),
            "--stored-pdf-path",
            str(pdf),
            "--market",
            "a_share",
        ]
    )
    report_id = capsys.readouterr().out.strip()

    class FakeExtractor:
        extractor_name = "pdfplumber"

        def extract_tables(self, pdf_path):
            assert pdf_path == pdf
            return [
                ExtractedTable(
                    extractor_name="pdfplumber",
                    page_number=1,
                    table_index_on_page=0,
                    bbox_json="[0, 0, 10, 10]",
                    cells=[
                        ExtractedCell(0, 0, "项目", None, 1),
                        ExtractedCell(0, 1, "金额", None, 1),
                    ],
                    quality={},
                )
            ]

    monkeypatch.setattr(
        "fin_report_extractor.cli.PdfPlumberExtractor",
        lambda: FakeExtractor(),
    )

    main(["extract-tables", report_id, "--audit-db", str(audit_db)])

    output = capsys.readouterr().out
    assert "extraction_run_id=" in output
    assert "tables=1" in output
    assert "cells=2" in output

    conn = sqlite3.connect(audit_db)
    try:
        run_count = conn.execute("select count(*) from extraction_runs").fetchone()[0]
        table_count = conn.execute("select count(*) from raw_tables").fetchone()[0]
        cell_count = conn.execute("select count(*) from raw_cells").fetchone()[0]
    finally:
        conn.close()

    assert run_count == 1
    assert table_count == 1
    assert cell_count == 2


def _setup_full_pipeline_db(tmp_path, monkeypatch):
    """Create a database with all three statements ready for export.

    Returns (audit_db_path, run_id).
    """
    from fin_report_extractor.audit_db import connect_audit_db, initialize_audit_db
    from fin_report_extractor.import_pdf import register_pdf

    audit_db = tmp_path / "db" / "audit.sqlite"
    audit_db.parent.mkdir(parents=True, exist_ok=True)
    conn = connect_audit_db(audit_db)
    initialize_audit_db(conn)

    pdf = tmp_path / "report.pdf"
    pdf.write_bytes(b"%PDF-1.4 report\n")

    report_id = register_pdf(
        conn, pdf, stored_pdf_path=str(pdf),
        market="a_share", company_id="000651", company_name="格力电器",
        fiscal_year=2025, report_type="quarterly",
    )

    for idx, (page, text) in enumerate([
        (1, "合并资产负债表\n单位：元"),
        (2, "合并利润表\n单位：元"),
        (3, "合并现金流量表\n单位：元"),
    ], start=1):
        conn.execute(
            """
            insert into pdf_pages (
              page_id, report_id, page_number, width, height, text_char_count,
              text_density, has_statement_keywords, page_text_sample
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (f"page-{idx}", report_id, page, 595.0, 842.0, len(text), 0.0002, 1, text),
        )
    conn.commit()

    # Build three-statement raw tables via persist_raw_tables
    from fin_report_extractor.extraction_runs import (
        create_extraction_run,
        persist_raw_tables,
    )
    run_id = create_extraction_run(conn, report_id, extractor_versions={"pdfplumber": "test"})

    def data_row(r, label, cur, prior=None, page=1):
        return [
            ExtractedCell(r, 0, None, None, page),
            ExtractedCell(r, 1, label, None, page),
            ExtractedCell(r, 2, None, None, page),
            ExtractedCell(r, 3, cur, None, page),
            ExtractedCell(r, 4, None, None, page),
            ExtractedCell(r, 5, None, None, page),
            ExtractedCell(r, 6, prior, None, page),
            ExtractedCell(r, 7, None, None, page),
            ExtractedCell(r, 8, None, None, page),
        ]

    bs = [
        ExtractedCell(0, 1, "项目", None, 1),
        ExtractedCell(0, 3, "期末余额", None, 1),
        ExtractedCell(0, 6, "期初余额", None, 1),
    ]
    bs.extend(data_row(1, "资产总计", "100", "90", 1))
    bs.extend(data_row(2, "负债合计", "40", "30", 1))
    bs.extend(data_row(3, "所有者权益合计", "60", "60", 1))

    is_cells = [
        ExtractedCell(0, 1, "项目", None, 2),
        ExtractedCell(0, 3, "本期发生额", None, 2),
        ExtractedCell(0, 6, "上期发生额", None, 2),
    ]
    is_cells.extend(data_row(1, "净利润", "100", "90", 2))
    is_cells.extend(data_row(2, "综合收益总额", "100", "90", 2))
    is_cells.extend(data_row(3, "基本每股收益", "1.00", "0.90", 2))

    cf = [
        ExtractedCell(0, 1, "项目", None, 3),
        ExtractedCell(0, 3, "本期发生额", None, 3),
        ExtractedCell(0, 6, "上期发生额", None, 3),
    ]
    cf.extend(data_row(1, "经营活动产生的现金流量净额", "100", None, 3))
    cf.extend(data_row(2, "投资活动产生的现金流量净额", "100", None, 3))
    cf.extend(data_row(3, "筹资活动产生的现金流量净额", "100", None, 3))
    cf.extend(data_row(4, "期末现金及现金等价物余额", "100", None, 3))

    persist_raw_tables(conn, report_id, run_id, [
        ExtractedTable("pdfplumber", 1, 0, None, bs, {}),
        ExtractedTable("pdfplumber", 2, 0, None, is_cells, {}),
        ExtractedTable("pdfplumber", 3, 0, None, cf, {}),
    ])

    from fin_report_extractor.table_classifier import classify_tables_for_run
    classify_tables_for_run(conn, run_id, rules_root=Path("rules"))
    conn.close()
    return audit_db, run_id


def test_export_statements_cli_writes_workbook(tmp_path, capsys, monkeypatch):
    audit_db, run_id = _setup_full_pipeline_db(tmp_path, monkeypatch)
    output = tmp_path / "output.xlsx"

    main([
        "export-statements", run_id,
        "--audit-db", str(audit_db),
        "--output", str(output),
    ])

    out = capsys.readouterr().out
    assert str(output) in out
    assert output.exists()

    from openpyxl import load_workbook
    wb = load_workbook(output)
    assert "资产负债表" in wb.sheetnames
    assert "利润表" in wb.sheetnames
    assert "现金流量表" in wb.sheetnames
    assert "说明" in wb.sheetnames


def test_export_pdf_statements_cli_runs_one_shot(tmp_path, capsys, monkeypatch):
    audit_db = tmp_path / "db" / "audit.sqlite"
    pdf = tmp_path / "report.pdf"
    pdf.write_bytes(b"%PDF-1.4 report\n")

    # Provide a fake extractor so the one-shot pipeline doesn't need
    # a real PDF that pdfplumber can parse.
    class FakeExtractor:
        extractor_name = "pdfplumber"

        def extract_tables(self, pdf_path):
            def dr(r, label, cur, prior=None, page=1):
                return [
                    ExtractedCell(r, 0, None, None, page),
                    ExtractedCell(r, 1, label, None, page),
                    ExtractedCell(r, 2, None, None, page),
                    ExtractedCell(r, 3, cur, None, page),
                    ExtractedCell(r, 4, None, None, page),
                    ExtractedCell(r, 5, None, None, page),
                    ExtractedCell(r, 6, prior, None, page),
                    ExtractedCell(r, 7, None, None, page),
                    ExtractedCell(r, 8, None, None, page),
                ]

            bs = [
                ExtractedCell(0, 1, "项目", None, 1),
                ExtractedCell(0, 3, "期末余额", None, 1),
                ExtractedCell(0, 6, "期初余额", None, 1),
            ]
            bs.extend(dr(1, "资产总计", "100", "90", 1))
            bs.extend(dr(2, "负债合计", "40", "30", 1))
            bs.extend(dr(3, "所有者权益合计", "60", "60", 1))

            is_cells = [
                ExtractedCell(0, 1, "项目", None, 2),
                ExtractedCell(0, 3, "本期发生额", None, 2),
                ExtractedCell(0, 6, "上期发生额", None, 2),
            ]
            is_cells.extend(dr(1, "净利润", "100", "90", 2))
            is_cells.extend(dr(2, "综合收益总额", "100", "90", 2))
            is_cells.extend(dr(3, "基本每股收益", "1.00", "0.90", 2))

            cf = [
                ExtractedCell(0, 1, "项目", None, 3),
                ExtractedCell(0, 3, "本期发生额", None, 3),
                ExtractedCell(0, 6, "上期发生额", None, 3),
            ]
            cf.extend(dr(1, "经营活动产生的现金流量净额", "100", None, 3))
            cf.extend(dr(2, "投资活动产生的现金流量净额", "100", None, 3))
            cf.extend(dr(3, "筹资活动产生的现金流量净额", "100", None, 3))
            cf.extend(dr(4, "期末现金及现金等价物余额", "100", None, 3))

            return [
                ExtractedTable("pdfplumber", 1, 0, None, bs, {}),
                ExtractedTable("pdfplumber", 2, 0, None, is_cells, {}),
                ExtractedTable("pdfplumber", 3, 0, None, cf, {}),
            ]

    monkeypatch.setattr(
        "fin_report_extractor.cli.PdfPlumberExtractor",
        lambda: FakeExtractor(),
    )

    # The stub PDF is not a real PDF — skip PyMuPDF profiling.
    # Insert page text manually so the table classifier can work.
    from dataclasses import dataclass

    @dataclass(frozen=True)
    class FakeProfile:
        report_id: str
        page_count: int = 3
        is_text_pdf: bool = True
        unsupported_reason: str | None = None
        keyword_page_count: int = 3

    def fake_profile(conn, report_id):
        conn.execute("update reports set page_count=3, is_text_pdf=1 where report_id=?", (report_id,))
        conn.execute("delete from pdf_pages where report_id = ?", (report_id,))
        for idx, text in enumerate(
            ["合并资产负债表\n单位：元", "合并利润表\n单位：元", "合并现金流量表\n单位：元"],
            start=1,
        ):
            conn.execute(
                """
                insert into pdf_pages (
                  page_id, report_id, page_number, width, height, text_char_count,
                  text_density, has_statement_keywords, page_text_sample
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (f"page-{idx}", report_id, idx, 595.0, 842.0, len(text), 0.0002, 1, text),
            )
        conn.commit()
        return FakeProfile(report_id)

    monkeypatch.setattr(
        "fin_report_extractor.cli.profile_pdf_for_report",
        fake_profile,
    )

    output = tmp_path / "statements.xlsx"
    main([
        "export-pdf-statements", str(pdf),
        "--market", "a_share",
        "--company-id", "000651",
        "--company-name", "格力电器",
        "--fiscal-year", "2025",
        "--report-type", "quarterly",
        "--audit-db", str(audit_db),
        "--output", str(output),
    ])

    out = capsys.readouterr().out
    assert str(output) in out
    assert output.exists()

    from openpyxl import load_workbook
    wb = load_workbook(output)
    assert "资产负债表" in wb.sheetnames
    assert "利润表" in wb.sheetnames
    assert "现金流量表" in wb.sheetnames
